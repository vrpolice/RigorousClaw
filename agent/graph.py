import os
from typing import Annotated, Sequence, TypedDict, Literal
from langchain_core.messages import BaseMessage, HumanMessage, ToolMessage, AIMessage, SystemMessage
from langchain_openai import ChatOpenAI
from langgraph.graph import StateGraph, END, add_messages
from langgraph.prebuilt import ToolNode

from config import load_config

# Import our custom skills
from .tools.web_tools import tavily_search, jina_reader, system_cli_command
from .tools.crawler_tools import robust_web_crawler
from .memory import save_to_memory, recall_from_memory, auto_recall

# For Document Reading (pure Python, no Java needed):
from langchain_core.tools import tool
import os

@tool
def tika_parse(file_path: str) -> str:
    """Read the text content of a local document (pdf, docx, xlsx, txt, etc.) given its absolute path."""
    if not os.path.exists(file_path):
        return f"Error: File not found: {file_path}"
    
    ext = os.path.splitext(file_path)[1].lower()
    
    try:
        if ext == '.pdf':
            from PyPDF2 import PdfReader
            reader = PdfReader(file_path)
            text = ""
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
            return text if text.strip() else "(PDF contains no extractable text, may be image-based)"
        
        elif ext in ('.docx', '.doc'):
            from docx import Document
            doc = Document(file_path)
            return "\n".join(p.text for p in doc.paragraphs)
        
        elif ext in ('.xlsx', '.xls'):
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True)
            text = ""
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                text += f"--- Sheet: {sheet} ---\n"
                for row in ws.iter_rows(values_only=True):
                    text += "\t".join(str(cell) if cell is not None else "" for cell in row) + "\n"
            wb.close()
            return text
        
        elif ext in ('.txt', '.md', '.csv', '.json', '.html', '.xml', '.log', '.py', '.js'):
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()
        
        else:
            # Try reading as plain text
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()
    except Exception as e:
        return f"Error reading file {file_path}: {e}"

# Define the state of our agent
class AgentState(TypedDict):
    messages: Annotated[Sequence[BaseMessage], add_messages]

# Define all available strict tools
tools = [
    tavily_search, 
    jina_reader, 
    robust_web_crawler, 
    system_cli_command, 
    save_to_memory, 
    recall_from_memory,
    tika_parse
]

def call_model(state: AgentState):
    """The central routing node. Determines if a tool should be called or if we should respond directly.
       Initializes the model per-request so that Config changes (Model Name, API Key, Base URL) apply immediately.
    """
    config = load_config()
    api_key = config["llm"].get("openai_api_key") or os.getenv("OPENAI_API_KEY")
    base_url = config["llm"].get("openai_base_url") or os.getenv("OPENAI_BASE_URL", None)
    model_name = config["llm"].get("model_name") or os.getenv("MODEL_NAME", "gpt-4o")
    
    # Initialize the LLM dynamically
    if base_url:
        model = ChatOpenAI(model=model_name, base_url=base_url, api_key=api_key, temperature=0.2)
    else:
        model = ChatOpenAI(model=model_name, api_key=api_key, temperature=0.2)
        
    model_with_tools = model.bind_tools(tools)
    
    # Build the dynamic System Prompt from Config
    persona_name = config["agent"].get("name", "Agent")
    persona_role = config["agent"].get("role", "Assistant")
    persona_style = config["agent"].get("style", "Professional")
    user_name = config["user"].get("name", "User")
    custom_system = config["agent"].get("system_prompt", "")
    
    prompt_text = (
        f"You are '{persona_name}', acting as a '{persona_role}'.\n"
        f"Your conversational style should be: {persona_style}.\n"
        f"You are assisting a user named: '{user_name}'.\n\n"
        f"CORE DIRECTIVES:\n{custom_system}\n"
        f"Remember to use your tools when needed (tavily_search, jina_reader, recall_from_memory, etc.).\n"
        f"When the user tells you something important about themselves or their preferences, use save_to_memory to save it."
    )
    
    # Auto-recall relevant memories based on the user's latest message
    user_messages = [m for m in state["messages"] if isinstance(m, HumanMessage)]
    if user_messages:
        latest_user_msg = user_messages[-1].content
        recalled = auto_recall(latest_user_msg)
        if recalled:
            prompt_text += f"\n\n{recalled}"

    sys_prompt = SystemMessage(content=prompt_text)
    messages = [sys_prompt] + list(state["messages"])
    
    response = model_with_tools.invoke(messages)
    return {"messages": [response]}

def should_continue(state: AgentState) -> Literal["tools", "__end__"]:
    """Determine if we should execute a tool or end the turn."""
    messages = state["messages"]
    last_message = messages[-1]
    
    # If the LLM made a tool call, we route to the 'tools' node
    if last_message.tool_calls:
        return "tools"
    # Otherwise, we end and return the response to the user
    return "__end__"

# Define the Tool Executor Node
tool_node = ToolNode(tools)

# Build the LangGraph workflow
workflow = StateGraph(AgentState)

# Add nodes
workflow.add_node("agent", call_model)
workflow.add_node("tools", tool_node)

# Set entry point
workflow.set_entry_point("agent")

# Add conditional edges
workflow.add_conditional_edges(
    "agent",
    should_continue,
)

# After tools run, always return to agent to synthesize the final response
workflow.add_edge("tools", "agent")

# Compile the graph
app = workflow.compile()
